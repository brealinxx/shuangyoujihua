from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import sys
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
import matplotlib.colors as mcolors
import openpyxl
import os
from io import BytesIO
import numpy as np

class Window(QWidget):
    def __init__(self):
        super().__init__()
        plt.rcParams['font.sans-serif']=['SimHei'] # move SimHei.ttf file to /path/to/your/virtualenv/lib/pythonX.X/site-packages/matplotlib/mpl-data/fonts/ttf/
        self.setWindowTitle("科能高级技工学校“双优计划”评分表")
        self.setGeometry(100, 100, 400, 200)

        self.vLayout = QVBoxLayout()
        hLayout = QHBoxLayout()
        
        #* 选择文件按钮
        select_file_button = QPushButton("选择文件", self)
        select_file_button.setFixedHeight(40)
        select_file_button.clicked.connect(self.select_file_button_click)

        #* 输入框
        self.path_input = QLineEdit(self)
        self.path_input.setPlaceholderText("请点击选择文件或手动键入 Excel 文件")
        font = QFont()
        font.setItalic(True)
        font.setPointSize(12)
        self.path_input.setFont(font)
        self.debounce_timer = QTimer(self)
        self.debounce_timer.setSingleShot(True)
        self.debounce_timer.timeout.connect(self.on_path_input_confirmed)
        self.path_input.textChanged.connect(self.on_path_input_change)
       
        #* 生成按钮
        # todo 消息闪现（成果、失败）
        image_generate_button = QPushButton("生成", self)
        image_generate_button.setFixedHeight(40)
        # file_generate.setStyleSheet("background-color: green")
        image_generate_button.clicked.connect(self.image_generate_button_click)

        #* 导出按钮
        image_export_button = QPushButton("导出", self)
        image_export_button.setFixedHeight(40)
        # file_export.setStyleSheet("background-color: yellow")
        image_export_button.clicked.connect(self.image_export_button_click)

        self.status_label = QLabel(self)
        self.image_label = QLabel(self)  

        self.vLayout.addWidget(select_file_button)
        self.vLayout.addWidget(self.path_input)

        hLayout.addWidget(image_generate_button)  
        hLayout.addWidget(image_export_button) 
        self.vLayout.addLayout(hLayout)

        self.vLayout.addWidget(self.status_label)
        self.vLayout.addWidget(self.image_label)  
    
        self.setLayout(self.vLayout)

        self.image_generated = False 
        self.file_path = ""

    def select_file_button_click(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Excel文件 (*.xlsx)")
        self.file_path, _ = file_dialog.getOpenFileName(self, "选择文件", "", "Excel文件 (*.xlsx)")
        if os.path.isfile(self.file_path):
            self.path_input.setText(self.file_path)
        else:
            QMessageBox.warning(self, "警告", "文件路径错误，请选择一个有效的 Excel 文件", QMessageBox.StandardButton.Ok)

    def on_path_input_change(self):
        self.debounce_timer.start(1000)

    def on_path_input_confirmed(self):
        if os.path.isfile(self.path_input.text()):
            self.file_path = self.path_input.text()
        else:
            QMessageBox.warning(self, "警告", "文件路径错误，请选择一个有效的 Excel 文件", QMessageBox.StandardButton.Ok)
        # self.status_label.setText("文本框内容：" + self.path_input.text())
    
    def CreatePie(self, subplot,value,title):
        data = {'部分':['任务完成得分', '目标达成得分', '资金使用得分', '文档规范性得分', '项目执行力得分'],'占比':value}
        df = pd.DataFrame(data)
        norm = mcolors.Normalize(vmin=0, vmax=max(df['占比']))
        colors = [Window.ColorMapping.cmap(norm(val)) for val in df['占比']]
        patches, texts, autotexts = subplot.pie(df['占比'], labels=df['部分'],autopct=df['部分'], colors=colors) # modify .venv/lib/python3.12/site-packages/matplotlib/axes/_axes.py 3313 line
        plt.setp(texts, color='none')
        #subplot.set_title(title, loc='left',fontsize=80,color='white')
        plt.axis('equal')
    
    def CreateHBarCharts(self,subplot,value,set_ylim1,set_ylim2,set_yticks,bar_height,title):#! 对应数据
        # subplot.set_xlim(0, 60)  # 设置x轴范围
        lefts = np.arange(len(self.categories)) * set_yticks
        for i, (value, category) in enumerate(zip(value, self.categories)):
            bar = subplot.barh(lefts[i], value, height=bar_height, color=Window.ColorMapping.cmap(Window.ColorMapping.norm(value/100)))
        subplot.set_facecolor('none')
        subplot.set_ylim(set_ylim1, set_ylim2)
        subplot.set_yticks(lefts + 0.5)
        subplot.set_yticklabels(self.categories,fontsize=25, color='yellow')
        #subplot.set_title(title, loc='left', fontsize=60,color='white')
        subplot.tick_params(bottom=False, colors='white', labelsize=30)
        # ax_middle.tick_params(axis='y', which='both', left=False)  # 设置y轴刻度参数   
    
    def CreateBarCharts(self, subplot,value,set_xlim1,set_xlim2,set_xticks,bar_width,title,titleTrigger, x_labels = None):
        subplot.set_ylabel('百分比 (%)')
        subplot.set_ylim(0, 100) 
        bottoms = np.arange(len(value)) * set_xticks
        for i, (val, category) in enumerate(zip(value, x_labels)):
            bar = subplot.bar(bottoms[i], val, width=bar_width, color=Window.ColorMapping.cmap(Window.ColorMapping.norm(val/100)), edgecolor='none')
            # subplot.axhline(val, linestyle='--', color='gray') 
        subplot.set_xlim(set_xlim1, set_xlim2)
        if x_labels:
            subplot.set_xticks(bottoms)
            subplot.set_xticklabels(x_labels, rotation=45, fontsize=15, color='yellow')

        if titleTrigger:
            subplot.set_title(title, fontsize=40,color='white')
        subplot.tick_params(left=False, colors='white', labelsize=30)
        subplot.grid(axis='y') 
        subplot.set_facecolor('none')

    class ColorMapping:
        def ColorTrans(r,g,b,a):
            return (r/255,g/255,b/255,a)
        
        colors = [ColorTrans(211,12,18,1.000), ColorTrans(242,92,5,1.000), ColorTrans(242,206,27,1.000), ColorTrans(15,113,242,1), ColorTrans(13,242,5,1.000)]
        cmap = mcolors.ListedColormap(colors)
        bounds = [0, 0.6, 0.75, 0.9, 1, 1]
        norm = mcolors.BoundaryNorm(bounds, cmap.N)
    
    def image_generate_button_click(self):
        # todo: image generate here
        if not self.file_path:
            QMessageBox.warning(self, "警告", "请先「选择文件」", QMessageBox.StandardButton.Ok)
            return

        if self.file_path:
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.sheets = workbook['sheet']
            self.sheetTest = workbook['test']
            self.colors = [Window.ColorMapping.ColorTrans(211,12,18,1.000), Window.ColorMapping.ColorTrans(242,92,5,1.000), Window.ColorMapping.ColorTrans(242,206,27,1.000), Window.ColorMapping.ColorTrans(15,113,242,1), Window.ColorMapping.ColorTrans(13,242,5,1.000)]

            # 保存图表为 QPixmap
            buffer = BytesIO()
            fig, axs = plt.subplots(figsize=(3840 / 72, 12000 / 72),facecolor=(3/255, 32/255, 71/255))
            axs.axis('off') 
            rowsCount = 13
            columnsCount = 5
            gs = GridSpec(rowsCount, columnsCount)
            axs = [[fig.add_subplot(gs[i, j]) for j in range(columnsCount)] for i in range(rowsCount)]
            self.ratio = [35,35,15,5,15]
            self.categories = ['任务完成率', '目标达成度', '资源使用率', '文档规范性', '项目执行力']

            def GetExcelData(definedName):
                # 通过定义名称获取单元格对象
                o = workbook.defined_names[definedName]
                # 获取定义名称的范围
                cells = o.destinations

                # 从单元格对象中获取值
                sheet_name, cell_range = next(cells)
                sheet = workbook[sheet_name]
                return sheet[cell_range].value
            
            def GetIntegerCount(renwuwancheng,mubiaodacheng,zijinshiyong,wendangguifanxing,xiangmuzhixingli):
                return [GetExcelData(renwuwancheng),GetExcelData(mubiaodacheng),GetExcelData(zijinshiyong),GetExcelData(wendangguifanxing),GetExcelData(xiangmuzhixingli)]
            
            def GetIntegerPercentage(values):
                return [(b / m * 100) for b , m in zip(values,self.ratio)]

            #todo data mapping
            values1 = GetIntegerCount("党建任务完成得分","党建目标达成得分","党建资金使用得分","党建文档规范性得分","党建项目执行力得分")
            self.CreatePie(fig.add_subplot(gs[0:1, 0:1]),values1,"党建考核")

            values2 = GetIntegerCount("信息化建设任务完成得分","信息化建设目标达成得分","信息化建设资金使用得分","信息化建设文档规范性得分","信息化建设项目执行力得分")
            self.CreatePie(fig.add_subplot(gs[0, 4:5]),values2,"信息化建设考核")

            values3 = GetIntegerCount("立德树人任务完成得分","立德树人目标达成得分","立德树人资金使用得分","立德树人文档规范性得分","立德树人项目执行力得分")
            self.CreatePie(fig.add_subplot(gs[2, 0:1]),values3,"立德树人考核")

            values4 = GetIntegerCount("社会服务任务完成得分","社会服务目标达成得分","社会服务资金使用得分","社会服务文档规范性得分","社会服务项目执行力得分")
            self.CreatePie(fig.add_subplot(gs[2, 4:5]),values4,"社会服务能力考核")

            values5 = GetIntegerCount("整体任务完成得分","整体目标达成得分","整体资金使用得分","整体文档规范性得分","整体项目执行力得分")
            self.CreateHBarCharts(fig.add_subplot(gs[1, 1:4]),GetIntegerPercentage(values5),-1,2,3,1,'整体考核')

            values6 = GetIntegerCount("治理体系任务完成得分","治理体系目标达成得分","治理体系资金使用得分","治理体系文档规范性得分","治理体系项目执行力得分")
            self.CreateHBarCharts(fig.add_subplot(gs[3, 0:3]),GetIntegerPercentage(values6),-1,2,3,1,'治理体系考核')
            
            #! 3
            values7 = [GetExcelData('国际任务完成得分') , GetExcelData('国际目标达成得分'), GetExcelData('国际资金使用得分'), GetExcelData('国际文档规范性得分'), GetExcelData('国际项目执行力得分')]
            self.CreateBarCharts(fig.add_subplot(gs[3, 4]),GetIntegerPercentage(values7),-1,2,7,3,'国际交流合作考核',False,self.categories)
            values8 = [GetExcelData('智能任务完成得分'), GetExcelData('智能目标达成得分'), GetExcelData('智能资金使用得分'), GetExcelData('智能文档规范性得分'), GetExcelData('智能项目执行力得分')]
            self.CreateBarCharts(fig.add_subplot(gs[4, 0]),GetIntegerPercentage(values8),-1,2,7,3,'智能制造专业考核',False,self.categories)
            values9 = [GetExcelData('交通任务完成得分'), GetExcelData('交通目标达成得分'), GetExcelData('交通资金使用得分'), GetExcelData('交通文档规范性得分'), GetExcelData('交通项目执行力得分')]
            self.CreateBarCharts(fig.add_subplot(gs[4, 2]),GetIntegerPercentage(values9),-1,2,7,3,'新能源交通考核',False,self.categories)
            values10 = [GetExcelData('现代任务完成得分'), GetExcelData('现代目标达成得分'), GetExcelData('现代资金使用得分'), GetExcelData('现代文档规范性得分'), GetExcelData('现代项目执行力得分')]
            self.CreateBarCharts(fig.add_subplot(gs[4, 4]),GetIntegerPercentage(values10),-1,2,7,3,'现代服务业考核',False,self.categories)


            #! 比例问题
            def ConvertCube(values):
                return sum(([b / m for b, m in zip(values, self.ratio)])) / 5

            box_width = 0.35
            box_height = 0.2
            x_positions = [0, 0.35, 0.7] 
            y_position = 0.5
            reacName = ['治理体系','党建','国际交流合作','立德树人','社会服务','信息化建设','新能源交通','智能制造','现代服务业']
            percentages = [ConvertCube(values6), ConvertCube(values1), ConvertCube(values7),ConvertCube(values3),ConvertCube(values4),ConvertCube(values2),ConvertCube(values9),ConvertCube(values8),ConvertCube(values10)]

            def CreateRectCharts(subplot, box_widthReduce ,textXPos, namePos):
                for i in range(3):
                    subplot.set_facecolor('none')
                    subplot.axis('off')
                    x = x_positions[i]
                    y = y_position
                    rect = plt.Rectangle((x, y), box_width - box_widthReduce, box_height, facecolor=Window.ColorMapping.cmap(Window.ColorMapping.norm(percentages[i + namePos])))
                    subplot.add_patch(rect)
                    subplot.text(x + box_width / 2 - textXPos , y + box_height / 2, f'{reacName[i + namePos]}', ha='center', va='center', fontsize=25)
                
            #! 4
            CreateRectCharts(fig.add_subplot(gs[5, 0:1]),0.07,0.04,0)
            CreateRectCharts(fig.add_subplot(gs[5, 4:5]),0.07,0.04,6)
            CreateRectCharts(fig.add_subplot(gs[5, 2:3]),0.07,0.04,3)
            
            #! 5
            row_start = 2
            resultVals = []
            self.names = []
            for row_num in range(row_start, self.sheetTest.max_row + 1):
                cell_value = self.sheetTest.cell(row=row_num, column=8).value
                if cell_value is None:
                    break
                resultVals.append(cell_value)
                self.names.append(self.sheetTest.cell(row=row_num, column=1).value)
            resultVals = [float(val) for val in resultVals]
            name_val_dict = dict(zip(self.names, resultVals))
            sorted_dict = dict(sorted(name_val_dict.items(), key=lambda item: item[1]))
            self.CreateBarCharts(fig.add_subplot(gs[6, :]), list(sorted_dict.values()), -1, 2, 7,3, '牵头人考核', False, list(sorted_dict.keys()))
            
            #! 6
            # in export method          

            #! 7 调整位置
            sheet = fig.add_subplot(gs[10, :])
            tasks_and_scores = {}
            for row in self.sheets.iter_rows(min_row=4, values_only=True): 
                name, task, score_1, score_2, score_3, score_4, score_5 = row[5], row[4], row[13], row[15], row[17], row[19], row[21]
                score_1 = 0 if score_1 is None else score_1
                score_2 = 0 if score_2 is None else score_2
                score_3 = 0 if score_3 is None else score_3
                score_4 = 0 if score_4 is None else score_4
                score_5 = 0 if score_5 is None else score_5
                score = round((score_1 + score_2 + score_3 + score_4 + score_5) / 100.0, 1)
                if task not in tasks_and_scores:
                    tasks_and_scores[task] = {name: score}  
                else:
                    tasks_and_scores[task][name] = score 

            flattened_data = []
            for task, names_scores in tasks_and_scores.items():
                for name, score in names_scores.items():
                    flattened_data.append((task, name, score))
            sorted_data = sorted(flattened_data, key=lambda x: x[2])
            sorted_tasks_and_scores = {}
            for task, name, score in sorted_data:
                if task not in sorted_tasks_and_scores:
                    sorted_tasks_and_scores[task] = {name: score}
                else:
                    sorted_tasks_and_scores[task][name] = score

            
            data = [['评分', '任务', '姓名']]
            for task, names in sorted_tasks_and_scores.items():
                for name, score in names.items():
                    if task and name:  
                        data.append([float(score), task, name]) 
            sorted_data = sorted(flattened_data, key=lambda x: x[2])

            table = sheet.table(cellText=data, colWidths=[0.03,0.5,0.07], loc='center', cellLoc='center')

            table.scale(2, 2)
            sheet.set_xlim(0, 1)
            sheet.set_ylim(0, 2)
            table.auto_set_font_size(False)
            table.set_fontsize(20)           
            for i in range(1, len(data)): 
                table[i, 0].set_facecolor(Window.ColorMapping.cmap(Window.ColorMapping.norm(data[i][0])))
                table[i, 0].set_text_props(weight='bold', color='white')
                for j in range(len(data[0])): 
                    table[(i, j)].set_edgecolor('white')
                    table[(i, j)].set_height(0.08)
                    if j != 0: 
                        table[(i, j)].set_facecolor('none')
                        table[(i, j)].get_text().set_color('white')
                        
            

            #! test delete frame
            for i in range(rowsCount):
                for j in range(columnsCount):
                    axs[i][j].axis('off')
            sheet.axis('off')
            

            gs.update(wspace=.4, hspace=.5) # 调整整体间距
            plt.subplots_adjust(top=.93, bottom=.07, right=.9, left=.1, hspace=0, wspace=0)
            plt.savefig(buffer, format='png')
            buffer.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buffer.getvalue())
            buffer.close()

            self.image_label.setPixmap(pixmap)
            self.image_label.adjustSize()

            # 创建一个QWidget，将QLabel放置在该QWidget中
            chart_widget = QWidget()
            chart_layout = QVBoxLayout()
            chart_layout.addWidget(self.image_label)
            chart_widget.setLayout(chart_layout)

            # 创建一个QScrollArea
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)  # 设置QScrollArea自适应大小
            scroll_area.setWidget(chart_widget)  # 将QWidget设置为QScrollArea的widget
            self.vLayout.addWidget(scroll_area)

            # 显示图片
            self.status_label.setText("已成功生成！请点按「导出」")
            self.image_generated = True
        else:
            QMessageBox.warning(self, "警告", "文件路径为空，请选择一个有效的 Excel 文件", QMessageBox.StandardButton.Ok)
        
        
    def image_export_button_click(self):
        def resource_path(relative_path):
            """获取打包后的资源文件路径"""
            try:
                # PyInstaller创建的临时文件夹路径
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)
        
        if not self.image_generated: 
            QMessageBox.warning(self, "警告", "请先「生成」图片", QMessageBox.StandardButton.Ok)
            return

        file_dialog = QFileDialog()
        # first
        file_path, _ = file_dialog.getSaveFileName(self, "保存文件", "", "PNG文件 (*.png)")
                
        if file_path:
            pixmap = self.image_label.pixmap()
            bg_pixmap = QPixmap(resource_path('background_1.png')) #* background.png

            combined_pixmap = QPixmap(pixmap.size())
            combined_pixmap.fill(Qt.transparent)

            painter = QPainter(combined_pixmap)
            painter.drawPixmap(0, 0, pixmap)  
            painter.drawPixmap(0, 0, bg_pixmap)  
            painter.end()

            combined_pixmap.save(file_path) 
            self.status_label.setText("已导出到:" + file_path)
        else:
            return
        
        # second
        leader_file_path, _ = QFileDialog.getSaveFileName(self, "保存牵头人考核指标文件", "", "PNG文件 (*.png)")
        row_start = 2
        iScore = []
        jScore = []
        kScore = []
        lScore = []
        mScore = []
        for row_num in range(row_start, self.sheetTest.max_row + 1):
            cell_value = self.sheetTest.cell(row=row_num, column=9).value
            if cell_value is None:
                break
            iScore.append(self.sheetTest.cell(row=row_num, column=9).value)
            jScore.append(self.sheetTest.cell(row=row_num, column=10).value)
            kScore.append(self.sheetTest.cell(row=row_num, column=11).value)
            lScore.append(self.sheetTest.cell(row=row_num, column=12).value)
            mScore.append(self.sheetTest.cell(row=row_num, column=13).value)
        buffer_leaderPic = BytesIO()
        all_scores = [iScore, jScore, kScore, lScore, mScore] 
        charts_per_row = 3  

        gs1 = GridSpec(14, charts_per_row,wspace=0.5, hspace=0.99)
        fig1, axs1 = plt.subplots(figsize=(30, 100),facecolor=(3/255, 32/255, 71/255))
        axs1.axis('off') 

        for chart_num in range(41):
            row = chart_num // charts_per_row  
            col = chart_num % charts_per_row  

            ax = fig1.add_subplot(gs1[row, col])               

            value = [scores[chart_num] for scores in all_scores if chart_num < len(scores)]
            values = [(i/j) * 100 for i,j in zip(value,[35,35,15,5,10])]
            self.CreateBarCharts(ax, values, -1,2,7,3, self.names[chart_num], True, self.categories)

        plt.savefig(buffer_leaderPic, format='png')
        buffer_leaderPic.seek(0)
        self.pixmap1 = QPixmap()
        self.pixmap1.loadFromData(buffer_leaderPic.getvalue())

                
        bg_pixmap = QPixmap(resource_path('background_2.png'))
        second_pixmap = QPixmap(self.pixmap1.size())

        painter = QPainter(second_pixmap)
        painter.drawPixmap(0, 0, self.pixmap1)  
        painter.drawPixmap(0, 0, bg_pixmap) 
        painter.end()
        second_pixmap.save(leader_file_path)

        buffer_leaderPic.close()

        if leader_file_path:
            second_pixmap.save(leader_file_path)
            self.status_label.setText("第一张图已导出到: " + file_path + "\n第二张图已导出到: " + leader_file_path)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec())